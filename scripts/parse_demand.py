"""Parse the WGC Gold Demand Trends XLSX into demand.json.

Tuned for the actual WGC layout (verified against GDT_Tables_Q126_EN.xlsx):

  Sheet "Gold Balance"  — annual cols 2..17 (2010..2025) + quarterly
                           cols 22..86 (Q1'10..Q1'26), header row 4.
                           Category rows include "Jewellery Fabrication",
                           "Technology", "Total Bar and Coin",
                           "ETFs and Similar Products", "Central Bank and
                           Other Institutions" — these map onto the
                           dashboard's 5 demand categories.

  Sheet "Jewellery"     — same column layout, rows = countries +
                           region roll-ups (Greater China, Middle East,
                           Americas, Europe ex CIS) and "Total above".
                           We drop the roll-ups so countries don't
                           double-count.

  Sheet "Bar and Coin"  — identical to Jewellery, country-level
                           retail-investment demand.

Output: data/parsed/demand.json with the shape consumed by the Demand
tab. If the file can't be parsed for any reason, writes a stub so the
dashboard renders an "updating" state instead of crashing.
"""
from __future__ import annotations

import json
import math
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data" / "parsed"

CATEGORY_KEYS = [
    "jewellery",
    "bar_and_coin",
    "etf",
    "central_banks",
    "technology",
]

# Gold-Balance row labels that match each dashboard category. Matched
# case-insensitively after lstrip(); first hit wins per category.
CATEGORY_ROW_HINTS: dict[str, list[str]] = {
    "jewellery":     ["jewellery fabrication", "jewellery demand", "jewellery"],
    "technology":    ["technology"],
    "bar_and_coin":  ["total bar and coin", "bar and coin"],
    "etf":           ["etfs and similar products", "etf"],
    "central_banks": ["central bank and other institutions", "central bank"],
}

# Supply-side rows from Gold Balance — these tell the supply story
# (mine output ceiling, recycled-gold spikes at price tops, producer
# hedging when miners turn bearish).
SUPPLY_KEYS = ["mine_production", "recycled_gold", "net_producer_hedging", "total_supply"]
SUPPLY_ROW_HINTS: dict[str, list[str]] = {
    "mine_production":      ["mine production"],
    "recycled_gold":        ["recycled gold"],
    "net_producer_hedging": ["net producer hedging"],
    "total_supply":         ["total supply"],
}

# Gold Prices sheet — taxonomy labels → compact keys + display unit.
PRICE_CURRENCIES: list[tuple[str, str, str, str]] = [
    # (key, taxonomy_substring, display_label, unit_text)
    ("usd_oz",  "us$/oz",  "USD",  "$/oz"),
    ("eur_oz",  "€/oz",    "EUR",  "€/oz"),
    ("gbp_oz",  "£/oz",    "GBP",  "£/oz"),
    ("chf_kg",  "chf/kg",  "CHF",  "CHF/kg"),
    ("jpy_g",   "¥/g",     "JPY",  "¥/g"),
    ("inr_10g", "rs/10g",  "INR",  "₹/10g"),
    ("rmb_g",   "rmb/g",   "CNY",  "¥/g"),
    ("try_g",   "tl/g",    "TRY",  "₺/g"),
]

# Country-sheet rows to drop (region roll-ups, residual buckets, and
# grand totals — would double-count against the leaf countries below).
EXCLUDE_LABELS = {
    "greater china",
    "middle east",
    "americas",
    "europe ex cis",
    "total above",
    "world total",
    "other & stock change",
    "other and stock change",
}

# Quarterly header parser: matches "Q1'26", "Q3 '22", "Q4'10" etc.
QUARTER_RX = re.compile(r"Q([1-4])\s*'?\s*(\d{2})", re.IGNORECASE)


def num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
        return round(f, 4)
    return None


def parse_quarter(cell: Any) -> str | None:
    """\"Q1'26\" → \"2026Q1\" (returns None if not a quarter cell)."""
    if not isinstance(cell, str):
        return None
    m = QUARTER_RX.search(cell)
    if not m:
        return None
    yy = int(m.group(2))
    year = 2000 + yy if yy < 80 else 1900 + yy
    return f"{year}Q{m.group(1)}"


def parse_year(cell: Any) -> str | None:
    """A 4-digit year header cell → \"2024\"."""
    if isinstance(cell, (int, float)) and 2000 <= cell <= 2099:
        return str(int(cell))
    if isinstance(cell, str):
        m = re.match(r"^\s*(20\d{2})\s*$", cell)
        if m:
            return m.group(1)
    return None


def column_indices(header_row: tuple[Any, ...]) -> tuple[dict[int, str], dict[int, str]]:
    """Return (year_cols, quarter_cols) — each {col_idx: label}."""
    years: dict[int, str] = {}
    quarters: dict[int, str] = {}
    for c, cell in enumerate(header_row):
        y = parse_year(cell)
        if y:
            years[c] = y
            continue
        q = parse_quarter(cell)
        if q:
            quarters[c] = q
    return years, quarters


def find_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    """Locate the header row that has either year or quarter columns.
    WGC files use row 4 (0-indexed) but be defensive."""
    for i in range(min(12, len(rows))):
        years, quarters = column_indices(rows[i])
        if years or quarters:
            return i
    return None


def classify_category(label: str) -> str | None:
    norm = label.lstrip().lower().strip()
    for cat, hints in CATEGORY_ROW_HINTS.items():
        for h in hints:
            if norm == h or norm.startswith(h):
                return cat
    return None


# ────────────────────────────────────────────────────────────────────
# Gold Balance — quarterly + annual time series for the 5 categories
# ────────────────────────────────────────────────────────────────────
def parse_gold_balance(wb) -> tuple[list[dict], list[dict]]:
    """Return (quarters, annual) lists of {key: ..., demand_tonnes: {...}}."""
    if "Gold Balance" not in wb.sheetnames:
        return [], []
    ws = wb["Gold Balance"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], []
    year_cols, quarter_cols = column_indices(rows[header_idx])

    # Walk rows looking for category labels in col 1
    cat_rows: dict[str, tuple[Any, ...]] = {}
    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        label = row[1] if len(row) > 1 else None
        if not isinstance(label, str) or not label.strip():
            continue
        cat = classify_category(label)
        if cat and cat not in cat_rows:
            cat_rows[cat] = row

    def build(col_map: dict[int, str]) -> list[dict]:
        # period_key (year or "2010Q1") -> {category: tonnes}
        out: dict[str, dict[str, float | None]] = {}
        for col_idx, key in col_map.items():
            bucket = out.setdefault(key, {cat: None for cat in CATEGORY_KEYS})
            for cat, row in cat_rows.items():
                bucket[cat] = num(row[col_idx]) if col_idx < len(row) else None
        # Order chronologically
        def sort_key(k: str):
            if "Q" in k:
                y, q = k.split("Q")
                return (int(y), int(q))
            return (int(k), 0)
        return [{"key": k, "demand_tonnes": out[k]} for k in sorted(out.keys(), key=sort_key)]

    quarters_raw = build(quarter_cols)
    annual_raw = build(year_cols)
    quarters = [{"quarter": p["key"], "demand_tonnes": p["demand_tonnes"]} for p in quarters_raw]
    annual = [{"year": p["key"], "demand_tonnes": p["demand_tonnes"]} for p in annual_raw]
    return quarters, annual


# ────────────────────────────────────────────────────────────────────
# Jewellery / Bar and Coin — per-country annual tonnes
# ────────────────────────────────────────────────────────────────────
def parse_country_sheet(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return []
    year_cols, _ = column_indices(rows[header_idx])
    if not year_cols:
        return []

    out: list[dict] = []
    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        label = row[1] if len(row) > 1 else None
        if not isinstance(label, str) or not label.strip():
            continue
        norm = label.strip().lower()
        if norm in EXCLUDE_LABELS:
            continue
        annual: dict[str, float] = {}
        for c_idx, year in year_cols.items():
            v = num(row[c_idx]) if c_idx < len(row) else None
            if v is not None:
                annual[year] = v
        if not annual:
            continue
        out.append({"country": label.strip(), "annual_tonnes": annual})

    # Sort by most-recent-year value, biggest first
    if out:
        latest_year = max(
            (y for r in out for y in r["annual_tonnes"].keys()),
            default=None,
        )
        out.sort(
            key=lambda r: -(r["annual_tonnes"].get(latest_year, 0) if latest_year else 0)
        )
    return out


# ────────────────────────────────────────────────────────────────────
# Gold Balance — supply rows (mine, recycled, producer hedging, total)
# ────────────────────────────────────────────────────────────────────
def parse_supply(wb) -> tuple[list[dict], list[dict]]:
    if "Gold Balance" not in wb.sheetnames:
        return [], []
    ws = wb["Gold Balance"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], []
    year_cols, quarter_cols = column_indices(rows[header_idx])

    supply_rows: dict[str, tuple[Any, ...]] = {}
    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        label = row[1] if len(row) > 1 else None
        if not isinstance(label, str) or not label.strip():
            continue
        norm = label.lstrip().lower().strip()
        for key, hints in SUPPLY_ROW_HINTS.items():
            if key in supply_rows:
                continue
            for h in hints:
                if norm == h or norm.startswith(h):
                    supply_rows[key] = row
                    break

    def build(col_map: dict[int, str]) -> list[dict]:
        out: dict[str, dict[str, float | None]] = {}
        for col_idx, key in col_map.items():
            bucket = out.setdefault(key, {k: None for k in SUPPLY_KEYS})
            for sk, row in supply_rows.items():
                bucket[sk] = num(row[col_idx]) if col_idx < len(row) else None
        def sort_key(k: str):
            if "Q" in k:
                y, q = k.split("Q")
                return (int(y), int(q))
            return (int(k), 0)
        return [{"key": k, "tonnes": out[k]} for k in sorted(out.keys(), key=sort_key)]

    qs = build(quarter_cols)
    ans = build(year_cols)
    quarters = [{"quarter": p["key"], "tonnes": p["tonnes"]} for p in qs]
    annual = [{"year": p["key"], "tonnes": p["tonnes"]} for p in ans]
    return quarters, annual


# ────────────────────────────────────────────────────────────────────
# Gold Prices — multi-currency annual + quarterly prices
# ────────────────────────────────────────────────────────────────────
def parse_gold_prices(wb) -> dict | None:
    if "Gold Prices" not in wb.sheetnames:
        return None
    ws = wb["Gold Prices"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return None
    year_cols, quarter_cols = column_indices(rows[header_idx])

    # Build a {currency_key: row_tuple} map by matching the Taxonomy cell
    price_rows: dict[str, tuple[Any, ...]] = {}
    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        label = row[1] if len(row) > 1 else None
        if not isinstance(label, str):
            continue
        norm = label.lower().replace(" ", "")
        for key, sub, _, _ in PRICE_CURRENCIES:
            if key in price_rows:
                continue
            if sub.replace(" ", "") in norm:
                price_rows[key] = row

    def build(col_map: dict[int, str]) -> list[dict]:
        out: dict[str, dict[str, float | None]] = {}
        for col_idx, period_key in col_map.items():
            bucket = out.setdefault(period_key, {})
            for ck, _, _, _ in PRICE_CURRENCIES:
                row = price_rows.get(ck)
                bucket[ck] = num(row[col_idx]) if row is not None and col_idx < len(row) else None
        def sort_key(k: str):
            if "Q" in k:
                y, q = k.split("Q")
                return (int(y), int(q))
            return (int(k), 0)
        return [{"key": k, "prices": out[k]} for k in sorted(out.keys(), key=sort_key)]

    quarters = [{"quarter": p["key"], "prices": p["prices"]} for p in build(quarter_cols)]
    annual = [{"year": p["key"], "prices": p["prices"]} for p in build(year_cols)]
    return {
        "currencies": [
            {"key": k, "label": lbl, "unit": unit}
            for k, _, lbl, unit in PRICE_CURRENCIES
        ],
        "annual": annual,
        "quarters": quarters,
    }


# ────────────────────────────────────────────────────────────────────
# Consumer per Capita — country-level grams per capita
# ────────────────────────────────────────────────────────────────────
def parse_per_capita(wb) -> list[dict]:
    if "Consumer per Capita" not in wb.sheetnames:
        return []
    ws = wb["Consumer per Capita"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return []
    year_cols, _ = column_indices(rows[header_idx])
    if not year_cols:
        return []
    out: list[dict] = []
    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        label = row[1] if len(row) > 1 else None
        if not isinstance(label, str) or not label.strip():
            continue
        norm = label.strip().lower()
        if norm in EXCLUDE_LABELS:
            continue
        annual: dict[str, float] = {}
        for c_idx, year in year_cols.items():
            v = num(row[c_idx]) if c_idx < len(row) else None
            if v is not None:
                annual[year] = v
        if not annual:
            continue
        out.append({"country": label.strip(), "annual_grams": annual})
    if out:
        latest = max((y for r in out for y in r["annual_grams"].keys()), default=None)
        out.sort(key=lambda r: -(r["annual_grams"].get(latest, 0) if latest else 0))
    return out


# ────────────────────────────────────────────────────────────────────
# Glue
# ────────────────────────────────────────────────────────────────────
_QUARTER_RX = re.compile(r"Q([1-4])[_-]?(\d{2,4})", re.IGNORECASE)


def _quarter_key(p: Path) -> tuple[int, int]:
    """(year, quarter) so sort picks the newest by date — NOT alphabetical.

    WGC's compact Q126/Q425 notation means Q425 > Q126 character-wise,
    so a plain sorted() would pick Q4'25 over Q1'26 — i.e. ignore the
    quarter that just became newer.
    """
    m = _QUARTER_RX.search(p.name)
    if not m:
        return (0, 0)
    quarter = int(m.group(1))
    year_token = m.group(2)
    year = int(year_token) if len(year_token) == 4 else 2000 + int(year_token)
    return (year, quarter)


def latest_demand_xlsx() -> Path | None:
    candidates: list[Path] = []
    for pattern in ("Gold_Demand_*.xlsx", "*emand*.xlsx", "GDT_*.xlsx"):
        candidates.extend(RAW_DIR.glob(pattern))
    # Dedupe
    seen: set[Path] = set()
    unique: list[Path] = []
    for p in candidates:
        if p in seen:
            continue
        seen.add(p)
        unique.append(p)
    if not unique:
        return None
    # Sort by (year, quarter); fall back to mtime if quarter token is absent.
    unique.sort(key=lambda p: (_quarter_key(p), p.stat().st_mtime), reverse=True)
    return unique[0]


def write_stub(reason: str) -> None:
    """Soft-fail: preserve existing demand.json if it already has quarters.
    Stops a parser crash on a fresh raw drop from wiping good data."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "demand.json"
    if out.exists():
        try:
            existing = json.load(open(out, encoding="utf-8"))
            if existing.get("quarters"):
                print(
                    f"[parse-demand] preserving existing {out.relative_to(ROOT)} "
                    f"(as_of={existing.get('as_of_quarter')}, "
                    f"{len(existing['quarters'])} quarters) — {reason}",
                    file=sys.stderr,
                )
                return
        except Exception:
            pass
    payload = {
        "as_of_quarter": None,
        "as_of_note": reason,
        "categories": CATEGORY_KEYS,
        "quarters": [],
        "annual": [],
        "by_country_jewellery": [],
        "by_country_bar_and_coin": [],
        "supply": {"quarters": [], "annual": []},
        "gold_prices": None,
        "per_capita_grams": [],
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"), ensure_ascii=False)
    print(f"[parse-demand] wrote stub {out.relative_to(ROOT)} — {reason}")


def main() -> None:
    xlsx = latest_demand_xlsx()
    if xlsx is None:
        write_stub("No WGC demand XLSX in data/raw/ yet — upload one and re-run.")
        return

    print(f"[parse-demand] Source: {xlsx.name}")
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    print(f"[parse-demand] Sheets: {', '.join(wb.sheetnames)}")

    quarters, annual = parse_gold_balance(wb)
    jewellery = parse_country_sheet(wb, "Jewellery")
    barcoin = parse_country_sheet(wb, "Bar and Coin")
    supply_q, supply_a = parse_supply(wb)
    gold_prices = parse_gold_prices(wb)
    per_capita = parse_per_capita(wb)

    if not quarters and not jewellery and not barcoin:
        write_stub(
            f"XLSX {xlsx.name} present but no recognizable sheets — "
            "WGC may have restructured. Inspect Gold Balance / Jewellery / "
            "Bar and Coin headers."
        )
        return

    as_of_quarter = quarters[-1]["quarter"] if quarters else None
    payload = {
        "as_of_quarter": as_of_quarter,
        "source_file": xlsx.name,
        "categories": CATEGORY_KEYS,
        "quarters": quarters,
        "annual": annual,
        "by_country_jewellery": jewellery,
        "by_country_bar_and_coin": barcoin,
        "supply": {"quarters": supply_q, "annual": supply_a},
        "gold_prices": gold_prices,
        "per_capita_grams": per_capita,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "demand.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"), ensure_ascii=False, default=str)
    print(
        f"[parse-demand] wrote {out.relative_to(ROOT)} "
        f"({out.stat().st_size:,} bytes, "
        f"{len(quarters)} quarters, {len(annual)} years, "
        f"{len(jewellery)} jewellery / {len(barcoin)} bar-coin countries, "
        f"{len(supply_q)} supply quarters, "
        f"{'gold prices ✓' if gold_prices else 'gold prices ✗'}, "
        f"{len(per_capita)} per-capita rows)"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[parse-demand] ERROR: {e}", file=sys.stderr)
        write_stub(f"Parser crashed: {e}")
