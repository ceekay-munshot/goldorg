"""Parse WGC / IMF-IFS Central Bank gold reserves into cb.json.

We expect TWO XLSX files in data/raw/:

  * World_official_gold_holdings_*.xlsx  (or *Holdings*.xlsx)
      One sheet, side-by-side ranking layout (2 countries per row):
        col 0   rank             col 5  rank
        col 1   country          col 6  country
        col 2   tonnes (current) col 7  tonnes
        col 3   % of reserves    col 8  % of reserves
        col 4   as-of date       col 9  as-of date

  * Changes_latest_*.xlsx (or *Changes*.xlsx)
      Sheets:
        "Monthly"  — country × month CHANGES (deltas in tonnes) Jan 2002→
        "Annual"   — country × year CHANGES

We join the two: current level from the holdings snapshot, monthly + annual
deltas from changes. Historical monthly LEVELS are back-derived from the
current snapshot using:
    monthly_tonnes[t-1] = monthly_tonnes[t] - monthly_change[t]

Output schema (data/parsed/cb.json):

  {
    "as_of_month": "2026-03",
    "as_of_holdings_date": "2026-03-31",
    "countries": [
      {
        "country": "United States",
        "current_tonnes": 8133.46,
        "pct_of_reserves": 0.833,
        "as_of_date": "2026-03-31",
        "monthly_change": { "2002-01": 0.0, ..., "2026-04": 0.0 },
        "annual_change":  { "2002":    0.0, ..., "2026":    0.0 },
        "monthly_tonnes": { "2002-01": ..., ..., "2026-04": 8133.46 }
      },
      ...
    ]
  }
"""
from __future__ import annotations

import json
import math
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data" / "parsed"

# Roll-up labels to drop so they don't appear as "countries".
EXCLUDE_LABELS = {
    "world", "world total", "total", "total above",
    "advanced economies", "emerging economies", "emerging markets",
    "developing economies", "eurozone", "euro area", "europe",
    "other countries", "other", "all countries",
    "data as of", "source:", "notes:", "n/a", "memo",
}


def num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
        return round(f, 6)
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        try:
            f = float(s)
            if math.isnan(f) or math.isinf(f):
                return None
            return round(f, 6)
        except ValueError:
            return None
    return None


def to_date_iso(v: Any) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d %b %Y", "%B %Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date().isoformat()
            except ValueError:
                continue
    return None


def to_ym(v: Any) -> str | None:
    iso = to_date_iso(v)
    if iso:
        return iso[:7]
    if isinstance(v, str):
        m = re.match(r"^(\d{4})[-/](\d{1,2})", v.strip())
        if m:
            return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    return None


def to_year(v: Any) -> str | None:
    if isinstance(v, (int, float)) and 1990 <= float(v) <= 2100:
        return str(int(v))
    if isinstance(v, datetime):
        return str(v.year)
    if isinstance(v, date):
        return str(v.year)
    if isinstance(v, str):
        m = re.match(r"^(\d{4})$", v.strip())
        if m:
            return m.group(1)
    return None


def clean_display_name(name: str) -> str:
    """Strip footnote markers like '5)', '*', '†' from the display label
    while preserving the rest of the country name."""
    s = re.sub(r"\s*\d+\s*\)+\s*$", "", name)
    s = re.sub(r"[\*†‡§¹²³]+\s*$", "", s)
    return s.strip()


def normalize_country_key(name: str) -> str:
    """Lowercase, strip footnote markers + punctuation, collapse whitespace —
    for matching across the holdings + changes files. Examples:
      'Turkey5)' → 'turkey'
      'Turkey*'  → 'turkey'
      'China, P.R.: Mainland' → 'china p r mainland'
      'Russian Federation'    → 'russian federation'
    """
    s = unicodedata.normalize("NFKD", name)
    s = s.lower()
    # Strip footnote markers: digits followed by ')' or '*' or '†' or '¹' etc.
    s = re.sub(r"\s*\d+\s*\)+\s*$", "", s)
    s = re.sub(r"[\*†‡§¹²³]+", "", s)
    s = re.sub(r"[,\.\:\-\(\)\[\]]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MONTH_YEAR_RX = re.compile(
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[_\-\s]?(\d{4})",
    re.IGNORECASE,
)
_YYYY_MM_RX = re.compile(r"(20\d{2})[-_](\d{1,2})")


def _month_key(p: Path) -> tuple[int, int]:
    """(year, month) so newest by date wins — NOT alphabetical.

    WGC names files Central_Bank_Holdings_May2026.xlsx. Sorted
    alphabetically, May > Jun > Mar/Sep/Nov for the same year because
    M comes after J and S alphabetically, so the parser would pick a
    stale May file over a newer June file. This parses the month token.
    """
    m = _MONTH_YEAR_RX.search(p.name)
    if m:
        return (int(m.group(2)), _MONTHS[m.group(1).lower()[:3]])
    m = _YYYY_MM_RX.search(p.name)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return (0, 0)


def find_xlsx(*patterns: str) -> Path | None:
    """Newest XLSX matching any of the patterns, sorted by (year, month)
    embedded in the filename (NOT alphabetical), with mtime as tiebreaker.
    """
    hits: list[Path] = []
    for pat in patterns:
        hits.extend(RAW_DIR.glob(pat))
    if not hits:
        return None
    hits = list(dict.fromkeys(hits))
    hits.sort(key=lambda p: (_month_key(p), p.stat().st_mtime), reverse=True)
    return hits[0]


# ────────────────────────────────────────────────────────────────────
# Holdings file — current snapshot, side-by-side ranking
# ────────────────────────────────────────────────────────────────────
def parse_holdings(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows:
            # Left side: cols 0..4
            left = parse_holdings_block(r, 0)
            if left:
                out.append(left)
            # Right side: cols 5..9
            right = parse_holdings_block(r, 5)
            if right:
                out.append(right)
        if out:
            break  # first sheet with rows wins
    return out


def parse_holdings_block(row: tuple[Any, ...], start: int) -> dict | None:
    if start + 4 >= len(row):
        return None
    rank_cell = row[start]
    name_cell = row[start + 1]
    tonnes_cell = row[start + 2]
    pct_cell = row[start + 3]
    asof_cell = row[start + 4]
    # Rank should be a small int; country a non-empty string; tonnes a number
    if not isinstance(name_cell, str) or not name_cell.strip():
        return None
    rank_val = num(rank_cell)
    if rank_val is None or rank_val < 1 or rank_val > 250:
        return None
    tonnes_val = num(tonnes_cell)
    if tonnes_val is None:
        return None
    name = name_cell.strip()
    if normalize_country_key(name) in EXCLUDE_LABELS:
        return None
    # % of reserves: sometimes "1)" footnote ref for IMF — store None then
    pct_val = num(pct_cell)
    return {
        "country": name,
        "current_tonnes": round(tonnes_val, 4),
        "pct_of_reserves": pct_val,
        "as_of_date": to_date_iso(asof_cell),
    }


# ────────────────────────────────────────────────────────────────────
# Changes file — Monthly + Annual sheets
# ────────────────────────────────────────────────────────────────────
def find_changes_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str], str] | None:
    """Return (header_row_idx, {col: 'YYYY-MM' or 'YYYY'}, granularity)."""
    # Try monthly first (looks for YYYY-MM dates)
    for r_idx in range(min(20, len(rows))):
        cols: dict[int, str] = {}
        for c, cell in enumerate(rows[r_idx]):
            ym = to_ym(cell)
            if ym:
                cols[c] = ym
        if len(cols) >= 24:
            return (r_idx, cols, "monthly")
    # Try annual
    for r_idx in range(min(8, len(rows))):
        cols: dict[int, str] = {}
        for c, cell in enumerate(rows[r_idx]):
            y = to_year(cell)
            if y:
                cols[c] = y
        if len(cols) >= 5:
            return (r_idx, cols, "annual")
    return None


def parse_changes_sheet(ws) -> tuple[dict[str, dict[str, float]], str]:
    """Return ({country: {period: change}}, granularity).

    The Changes file's Monthly sheet has TWO name columns:
      col 0 = "Country Lookup Column"  e.g. "China, People's Republic of"
      col 1 = "Country"                e.g. "China, P.R.: Mainland"
    We prefer col 1 — it matches the Holdings file's naming.
    Annual sheet uses col 0 as the only country column.
    """
    rows = list(ws.iter_rows(values_only=True))
    found = find_changes_header_row(rows)
    if not found:
        return {}, ""
    header_idx, period_cols, gran = found

    out: dict[str, dict[str, float]] = {}
    for r in rows[header_idx + 1:]:
        country: str | None = None
        # Prefer col 1 (the cleaner "Country" name) when non-empty,
        # otherwise fall back to col 0 (the "Country Lookup" IFS name).
        for c in (1, 0, 2):
            if c < len(r) and isinstance(r[c], str) and r[c].strip():
                cand = r[c].strip()
                if normalize_country_key(cand) not in EXCLUDE_LABELS:
                    country = cand
                    break
        if not country:
            continue
        country_map: dict[str, float] = {}
        for col_idx, period in period_cols.items():
            v = num(r[col_idx]) if col_idx < len(r) else None
            if v is not None:
                country_map[period] = v
        if not country_map:
            continue
        # If we've already seen this normalized country, prefer the entry
        # WITHOUT a '*' suffix (WGC marks superseded series with '*').
        norm_key = normalize_country_key(country)
        existing = next(
            (k for k in out if normalize_country_key(k) == norm_key),
            None,
        )
        if existing:
            if "*" in existing and "*" not in country:
                out.pop(existing)
            else:
                continue
        out[country] = country_map
    return out, gran


def parse_changes_file(path: Path) -> tuple[dict[str, dict[str, float]], dict[str, dict[str, float]]]:
    """Returns (monthly_changes, annual_changes) each keyed by country."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    monthly: dict[str, dict[str, float]] = {}
    annual: dict[str, dict[str, float]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data, gran = parse_changes_sheet(ws)
        if gran == "monthly" and not monthly:
            monthly = data
        elif gran == "annual" and not annual:
            annual = data
    return monthly, annual


# ────────────────────────────────────────────────────────────────────
# Merge + back-derive historical levels
# ────────────────────────────────────────────────────────────────────
def derive_monthly_tonnes(
    current_tonnes: float,
    monthly_change: dict[str, float],
    anchor_month: str,
) -> dict[str, float]:
    """Walk backwards from the anchor month using monthly_change to
    reconstruct the level series.

    Convention: monthly_change[t] = level(t) − level(t − 1), so
    level(t − 1) = level(t) − monthly_change[t].
    """
    if not monthly_change:
        return {anchor_month: round(current_tonnes, 4)}
    months = sorted(monthly_change.keys())
    # Anchor must be >= the earliest change month; if it's later than the
    # latest change month, walk forward from the latest known.
    # Build forward from earliest backward to latest of (changes ∪ {anchor}).
    all_months = set(months) | {anchor_month}
    all_months_sorted = sorted(all_months)
    # Find the index of the anchor
    if anchor_month not in all_months_sorted:
        return {anchor_month: round(current_tonnes, 4)}
    anchor_idx = all_months_sorted.index(anchor_month)
    out: dict[str, float] = {anchor_month: round(current_tonnes, 4)}
    # Backward
    level = current_tonnes
    for i in range(anchor_idx, 0, -1):
        m = all_months_sorted[i]
        prev_m = all_months_sorted[i - 1]
        delta = monthly_change.get(m, 0.0)
        level = level - delta
        out[prev_m] = round(level, 4)
    # Forward (in case changes file goes a month past the holdings as-of)
    level = current_tonnes
    for i in range(anchor_idx + 1, len(all_months_sorted)):
        m = all_months_sorted[i]
        delta = monthly_change.get(m, 0.0)
        level = level + delta
        out[m] = round(level, 4)
    return out


def merge_country_data(
    holdings: list[dict],
    monthly_changes: dict[str, dict[str, float]],
    annual_changes: dict[str, dict[str, float]],
) -> list[dict]:
    # Build name-key lookups for the changes files
    mc_index = {normalize_country_key(c): c for c in monthly_changes}
    ac_index = {normalize_country_key(c): c for c in annual_changes}

    out: list[dict] = []
    for h in holdings:
        key = normalize_country_key(h["country"])
        mc_orig = mc_index.get(key)
        ac_orig = ac_index.get(key)
        mc = monthly_changes.get(mc_orig, {}) if mc_orig else {}
        ac = annual_changes.get(ac_orig, {}) if ac_orig else {}
        # anchor month for back-derivation: prefer the holdings as-of date
        as_of_iso = h.get("as_of_date")
        anchor_month: str
        if as_of_iso:
            anchor_month = as_of_iso[:7]
        elif mc:
            anchor_month = max(mc.keys())
        else:
            anchor_month = ""
        monthly_tonnes = (
            derive_monthly_tonnes(h["current_tonnes"], mc, anchor_month)
            if anchor_month
            else {}
        )
        out.append({
            "country": clean_display_name(h["country"]),
            "current_tonnes": h["current_tonnes"],
            "pct_of_reserves": h.get("pct_of_reserves"),
            "as_of_date": as_of_iso,
            "monthly_change": mc,
            "annual_change": ac,
            "monthly_tonnes": monthly_tonnes,
        })
    return out


def write_stub(reason: str) -> None:
    """Soft-fail: preserve existing cb.json if it already has countries.
    Stops a parser crash on a fresh raw drop from wiping good data."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "cb.json"
    if out.exists():
        try:
            existing = json.load(open(out, encoding="utf-8"))
            if existing.get("countries"):
                print(
                    f"[parse-cb] preserving existing {out.relative_to(ROOT)} "
                    f"({len(existing['countries'])} countries) — {reason}",
                    file=sys.stderr,
                )
                return
        except Exception:
            pass
    payload = {
        "as_of_month": None,
        "as_of_note": reason,
        "countries": [],
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"), ensure_ascii=False)
    print(f"[parse-cb] wrote stub {out.relative_to(ROOT)} — {reason}")


def main() -> None:
    holdings_path = find_xlsx(
        "*Holdings*.xlsx", "World_official*.xlsx", "Central_Bank_Holdings*.xlsx",
    )
    changes_path = find_xlsx(
        "*Changes*.xlsx", "Central_Bank_Changes*.xlsx",
    )

    if not holdings_path and not changes_path:
        write_stub(
            "No central-bank XLSX in data/raw/ yet. Download the Holdings + "
            "Changes files from gold.org/goldhub/data/monthly-central-bank-statistics."
        )
        return

    # Partial input is dangerous: if Holdings is present but Changes is
    # missing (or vice versa), the merged output would drop every
    # country's monthly_change history → dashboard panels go blank.
    # Preserve the existing cb.json on partial input instead.
    if not (holdings_path and changes_path):
        write_stub(
            f"Partial CB input — holdings={'yes' if holdings_path else 'no'}, "
            f"changes={'yes' if changes_path else 'no'}. Need both to "
            "rebuild monthly history without losing it. Preserving existing cb.json."
        )
        return

    print(f"[parse-cb] holdings: {holdings_path.name}")
    print(f"[parse-cb] changes:  {changes_path.name}")

    holdings = parse_holdings(holdings_path)
    monthly_changes, annual_changes = parse_changes_file(changes_path)
    print(
        f"[parse-cb] parsed: {len(holdings)} holdings, "
        f"{len(monthly_changes)} monthly-changes countries, "
        f"{len(annual_changes)} annual-changes countries"
    )

    if not holdings:
        write_stub(
            f"Holdings file {holdings_path.name} didn't yield any countries — "
            "sheet layout may have changed. Preserving existing cb.json."
        )
        return

    if not monthly_changes:
        write_stub(
            f"Changes file {changes_path.name} didn't yield any monthly data "
            "— sheet layout may have changed. Preserving existing cb.json."
        )
        return

    countries = merge_country_data(holdings, monthly_changes, annual_changes)
    countries.sort(key=lambda c: -(c.get("current_tonnes") or 0))

    all_months: list[str] = []
    for c in countries:
        all_months.extend(c["monthly_change"].keys())
        all_months.extend(c["monthly_tonnes"].keys())
    as_of_month = max(all_months) if all_months else None

    as_of_dates = [c["as_of_date"] for c in countries if c.get("as_of_date")]
    as_of_holdings = max(as_of_dates) if as_of_dates else None

    payload = {
        "as_of_month": as_of_month,
        "as_of_holdings_date": as_of_holdings,
        "source_holdings": holdings_path.name,
        "source_changes": changes_path.name,
        "countries": countries,
    }

    # Atomic write: tmp + replace so a crash mid-dump can't truncate
    # a working cb.json that the dashboard depends on.
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "cb.json"
    tmp = out.with_suffix(out.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"), ensure_ascii=False, default=str)
        f.flush()
        try:
            import os
            os.fsync(f.fileno())
        except OSError:
            pass
    tmp.replace(out)
    print(
        f"[parse-cb] wrote {out.relative_to(ROOT)} "
        f"({out.stat().st_size:,} bytes, {len(countries)} countries, "
        f"as_of={as_of_month}, holdings_as_of={as_of_holdings})"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[parse-cb] ERROR: {e}", file=sys.stderr)
        write_stub(f"Parser crashed: {e}")
