"""Parse the gold.org ETF Flows XLSX into dashboard-ready JSON files.

Outputs (under data/parsed/):
  - metadata.json       file info, as-of, period date ranges
  - funds.json          per-fund snapshot + 7-period metrics
  - regions.json        4 regions + global, aggregated for 7 periods
  - countries.json      country-level aggregates for 7 periods
  - top_movers.json     top/bottom 15 by flow for each period
  - timeseries.json     monthly + annual regional series (full history)
  - fund_history.json   per-fund monthly history for drilldown overlay

Periods supported: 1M, QTD, YTD, 1Y, 3Y, 5Y, Max
"""
from __future__ import annotations

import argparse
import json
import math
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data" / "parsed"

PERIODS = ["1M", "QTD", "YTD", "1Y", "3Y", "5Y", "Max"]


# ============================================================
# Helpers
# ============================================================
def latest_xlsx() -> Path:
    files = sorted(RAW_DIR.glob("ETF_Flows_*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No ETF_Flows_*.xlsx found in {RAW_DIR}")
    return files[-1]


def num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
        return round(f, 6)
    return None


def to_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def sheet_rows(wb, name: str) -> list[list[Any]]:
    return [list(r) for r in wb[name].iter_rows(values_only=True)]


def write_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, separators=(",", ":"), ensure_ascii=False, default=str)
    print(f"[parse] wrote {path.relative_to(ROOT)} ({path.stat().st_size:,} bytes)")


# ============================================================
# Period definitions (relative to as-of date)
# ============================================================
def period_window(as_of: date, period: str) -> tuple[date, date]:
    """Return (from_exclusive, to_inclusive) for a period.
    Flows/demand are summed for months t where from < month_end <= to."""
    if period == "1M":
        # The single most recent month (e.g. Apr → from = end of Mar, to = end of Apr)
        prev = (as_of.replace(day=1) - timedelta(days=1))
        return prev, as_of
    if period == "QTD":
        # quarter to date
        q_start_month = ((as_of.month - 1) // 3) * 3 + 1
        q_start = date(as_of.year, q_start_month, 1)
        # Window starts at end of month before q_start (so quarter's first month is included)
        prev = q_start - timedelta(days=1)
        return prev, as_of
    if period == "YTD":
        prev = date(as_of.year - 1, 12, 31)
        return prev, as_of
    if period == "1Y":
        # trailing 12 months
        try:
            start = date(as_of.year - 1, as_of.month, as_of.day)
        except ValueError:
            start = date(as_of.year - 1, as_of.month, 28)
        return start, as_of
    if period == "3Y":
        return date(as_of.year - 3, as_of.month, min(as_of.day, 28)), as_of
    if period == "5Y":
        return date(as_of.year - 5, as_of.month, min(as_of.day, 28)), as_of
    if period == "Max":
        return date(1900, 1, 1), as_of
    raise ValueError(period)


def period_label(period: str, as_of: date) -> str:
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    if period == "1M":
        return f"{months[as_of.month - 1]} {as_of.year % 100:02d}"
    if period == "QTD":
        q = (as_of.month - 1) // 3 + 1
        return f"Q{q} {as_of.year % 100:02d} (QTD)"
    if period == "YTD":
        return f"YTD {as_of.year % 100:02d}"
    if period == "1Y":
        return "Trailing 1Y"
    if period == "3Y":
        return "Trailing 3Y"
    if period == "5Y":
        return "Trailing 5Y"
    if period == "Max":
        return "Since 2003"
    return period


# ============================================================
# Fund-level extraction from the 3 wide monthly sheets
# ============================================================
def extract_fund_universe(wb) -> tuple[list[dict], list[date], dict[str, list[float | None]], dict[str, list[float | None]], dict[str, list[float | None]]]:
    """Read the three wide monthly sheets and return:
      - funds_meta: [{ticker, name, region, country, active, fund_type, col_idx}]
      - dates: ordered list of month-end dates
      - holdings/demand/flows: {ticker: [monthly values]}
    """
    sheet_holdings = wb["Holdings by month"]
    sheet_demand = wb["Demand by month"]
    sheet_flows = wb["Fund flows by month"]

    h_rows = list(sheet_holdings.iter_rows(values_only=True))
    d_rows = list(sheet_demand.iter_rows(values_only=True))
    f_rows = list(sheet_flows.iter_rows(values_only=True))

    # Header structure (0-indexed):
    # row 0: ticker (col 5+); row 1: Active/Inactive; row 2: Fund Type
    # row 3: Region; row 4: Country; row 5: Fund Name + date col header
    # Data starts row 6.
    funds_meta = []
    n_cols = sheet_holdings.max_column
    for c in range(5, n_cols):
        ticker = h_rows[0][c]
        if not ticker or not isinstance(ticker, str):
            continue
        funds_meta.append(
            {
                "col_idx": c,
                "ticker": str(ticker).strip(),
                "active": str(h_rows[1][c]).strip().lower() == "active",
                "fund_type": str(h_rows[2][c]).strip() if h_rows[2][c] else None,
                "region": str(h_rows[3][c]).strip() if h_rows[3][c] else None,
                "country": str(h_rows[4][c]).strip() if h_rows[4][c] else None,
                "name": str(h_rows[5][c]).strip() if h_rows[5][c] else None,
            }
        )

    dates: list[date] = []
    holdings: dict[str, list[float | None]] = {f["ticker"]: [] for f in funds_meta}
    demand: dict[str, list[float | None]] = {f["ticker"]: [] for f in funds_meta}
    flows: dict[str, list[float | None]] = {f["ticker"]: [] for f in funds_meta}

    for hr, dr, fr in zip(h_rows[6:], d_rows[6:], f_rows[6:]):
        d = to_date(hr[0])
        if d is None:
            continue
        dates.append(d)
        for f in funds_meta:
            c = f["col_idx"]
            holdings[f["ticker"]].append(num(hr[c]))
            demand[f["ticker"]].append(num(dr[c]))
            flows[f["ticker"]].append(num(fr[c]))

    return funds_meta, dates, holdings, demand, flows


def global_series(wb) -> dict:
    """Extract the global aggregate columns (B-E + first 5 cols) from Holdings sheet."""
    rows = list(wb["Holdings by month"].iter_rows(values_only=True))
    out = {"dates": [], "gold_price_usd_oz": [], "total_ounces": [], "total_tonnes": [], "total_aum_usd": []}
    for r in rows[6:]:
        d = to_date(r[0])
        if d is None:
            continue
        out["dates"].append(d.isoformat())
        out["gold_price_usd_oz"].append(num(r[1]))
        out["total_ounces"].append(num(r[2]))
        out["total_tonnes"].append(num(r[3]))
        out["total_aum_usd"].append(num(r[4]))
    return out


# ============================================================
# Period metric computation
# ============================================================
def sum_window(values: list[float | None], dates: list[date], window_from: date, window_to: date) -> float:
    """Sum values for months where window_from < month_end <= window_to. None -> 0."""
    total = 0.0
    for v, d in zip(values, dates):
        if window_from < d <= window_to and v is not None:
            total += v
    return total


def holdings_at(values: list[float | None], dates: list[date], target: date) -> float | None:
    """Holdings value at the date <= target. Returns None if none found."""
    last = None
    for v, d in zip(values, dates):
        if d <= target and v is not None:
            last = v
    return last


def compute_fund_periods(
    meta: dict,
    dates: list[date],
    holdings: list[float | None],
    demand: list[float | None],
    flows: list[float | None],
    gold_price: list[float | None],
    as_of: date,
) -> dict:
    """Return dict {period: {metrics}} for a single fund."""
    out: dict[str, dict] = {}
    # current aum derived from holdings * gold price (tonnes -> oz -> usd)
    cur_holdings = holdings_at(holdings, dates, as_of)
    cur_price = holdings_at(gold_price, dates, as_of)
    cur_aum_usd_mn = None
    if cur_holdings is not None and cur_price is not None:
        # 1 tonne = 32,150.7466 troy oz
        cur_aum_usd_mn = round(cur_holdings * 32150.7466 * cur_price / 1e6, 4)

    for period in PERIODS:
        wf, wt = period_window(as_of, period)
        flows_sum = sum_window(flows, dates, wf, wt)
        demand_sum = sum_window(demand, dates, wf, wt)
        h_start = holdings_at(holdings, dates, wf)
        h_end = holdings_at(holdings, dates, wt)
        holdings_change = (h_end - h_start) if (h_start is not None and h_end is not None) else None
        demand_pct = None
        if cur_holdings and cur_holdings != 0:
            demand_pct = round(demand_sum / cur_holdings, 6)
        out[period] = {
            "flows_usd_mn": round(flows_sum, 4),
            "demand_tonnes": round(demand_sum, 6),
            "holdings_change_tonnes": round(holdings_change, 6) if holdings_change is not None else None,
            "demand_pct_of_holdings": demand_pct,
        }
    return {
        "current_holdings_tonnes": cur_holdings,
        "current_aum_usd_mn": cur_aum_usd_mn,
        "periods": out,
    }


# ============================================================
# Top-level parse functions
# ============================================================
def parse_metadata(wb, src: Path, as_of: date) -> dict:
    period_meta = {}
    for p in PERIODS:
        wf, wt = period_window(as_of, p)
        period_meta[p] = {
            "label": period_label(p, as_of),
            "from": wf.isoformat(),
            "to": wt.isoformat(),
        }
    return {
        "source_file": src.name,
        "source_size_bytes": src.stat().st_size,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "as_of_date": as_of.isoformat(),
        "periods": period_meta,
    }


def parse_funds(meta_list, dates, holdings, demand, flows, gp, as_of) -> dict:
    funds = []
    for m in meta_list:
        per = compute_fund_periods(
            m, dates, holdings[m["ticker"]], demand[m["ticker"]], flows[m["ticker"]], gp, as_of
        )
        # First and last months where this fund reported holdings (>0)
        first_active = None
        last_active = None
        for d, h in zip(dates, holdings[m["ticker"]]):
            if h is not None and h > 0:
                first_active = d
                break
        for d, h in zip(reversed(dates), reversed(holdings[m["ticker"]])):
            if h is not None and h > 0:
                last_active = d
                break
        # Last 36 months of monthly flows (USD mn) — used for streak strips
        flows_recent_36 = [
            (round(v, 4) if v is not None else None) for v in flows[m["ticker"]][-36:]
        ]
        # Skip funds with no holdings AND no flows (inactive + zero history)
        if per["current_holdings_tonnes"] in (None, 0) and all(
            per["periods"][p]["flows_usd_mn"] == 0 for p in PERIODS
        ):
            continue
        funds.append(
            {
                "ticker": m["ticker"],
                "name": m["name"],
                "region": m["region"],
                "country": m["country"],
                "active": m["active"],
                "fund_type": m["fund_type"],
                "first_active_date": first_active.isoformat() if first_active else None,
                "last_active_date": last_active.isoformat() if last_active else None,
                "current_holdings_tonnes": per["current_holdings_tonnes"],
                "current_aum_usd_mn": per["current_aum_usd_mn"],
                "periods": per["periods"],
                "flows_recent_36m": flows_recent_36,
            }
        )
    return {"count": len(funds), "funds": funds}


def aggregate(funds: list[dict], key: str) -> dict:
    """Aggregate fund period metrics by region or country."""
    buckets: dict[str, dict] = defaultdict(
        lambda: {
            "current_holdings_tonnes": 0.0,
            "current_aum_usd_mn": 0.0,
            "fund_count": 0,
            "periods": {p: {"flows_usd_mn": 0.0, "demand_tonnes": 0.0} for p in PERIODS},
        }
    )
    for f in funds:
        k = f.get(key) or "Unknown"
        b = buckets[k]
        if f["current_holdings_tonnes"]:
            b["current_holdings_tonnes"] += f["current_holdings_tonnes"]
        if f["current_aum_usd_mn"]:
            b["current_aum_usd_mn"] += f["current_aum_usd_mn"]
        b["fund_count"] += 1
        for p in PERIODS:
            pm = f["periods"][p]
            b["periods"][p]["flows_usd_mn"] += pm["flows_usd_mn"] or 0
            b["periods"][p]["demand_tonnes"] += pm["demand_tonnes"] or 0

    rows = []
    for name, b in buckets.items():
        # rounding pass
        for p in PERIODS:
            b["periods"][p]["flows_usd_mn"] = round(b["periods"][p]["flows_usd_mn"], 4)
            b["periods"][p]["demand_tonnes"] = round(b["periods"][p]["demand_tonnes"], 6)
            # add demand_pct
            ch = b["current_holdings_tonnes"]
            b["periods"][p]["demand_pct_of_holdings"] = (
                round(b["periods"][p]["demand_tonnes"] / ch, 6) if ch else None
            )
        b["current_holdings_tonnes"] = round(b["current_holdings_tonnes"], 6)
        b["current_aum_usd_mn"] = round(b["current_aum_usd_mn"], 4)
        rows.append({key: name, **b})
    # sort by current AUM desc
    rows.sort(key=lambda r: r["current_aum_usd_mn"] or 0, reverse=True)
    return {"count": len(rows), key + "s": rows}


def parse_top_movers(funds: list[dict], n: int = 15) -> dict:
    out: dict[str, dict] = {}
    for p in PERIODS:
        sorted_in = sorted(funds, key=lambda f: f["periods"][p]["flows_usd_mn"] or 0, reverse=True)
        sorted_out = sorted(funds, key=lambda f: f["periods"][p]["flows_usd_mn"] or 0)
        sorted_demand_pct_top = sorted(
            funds, key=lambda f: f["periods"][p]["demand_pct_of_holdings"] or 0, reverse=True
        )
        sorted_demand_pct_bot = sorted(
            funds, key=lambda f: f["periods"][p]["demand_pct_of_holdings"] or 0
        )

        def project(fs, count):
            return [
                {
                    "ticker": f["ticker"],
                    "name": f["name"],
                    "country": f["country"],
                    "region": f["region"],
                    "flows_usd_mn": f["periods"][p]["flows_usd_mn"],
                    "demand_tonnes": f["periods"][p]["demand_tonnes"],
                    "demand_pct_of_holdings": f["periods"][p]["demand_pct_of_holdings"],
                    "current_holdings_tonnes": f["current_holdings_tonnes"],
                }
                for f in fs[:count]
            ]

        out[p] = {
            "top_flows": project(sorted_in, n),
            "bottom_flows": project(sorted_out, n),
            "top_demand_pct": project(sorted_demand_pct_top, n),
            "bottom_demand_pct": project(sorted_demand_pct_bot, n),
        }
    return out


def parse_timeseries(wb) -> dict:
    """Re-extract pre-built chart series + add global aggregate series."""
    rows = list(wb["Charts Data"].iter_rows(values_only=True))
    blocks = {
        "monthly_flows_usd": (0, "gold_price_usd_oz"),
        "monthly_demand_tonnes": (9, "gold_price_usd_oz"),
        "annual_flows_usd": (18, "total"),
        "annual_demand_tonnes": (26, "total"),
        "monthly_holdings_tonnes": (34, "gold_price_usd_oz"),
        "monthly_holdings_usd": (43, "gold_price_usd_oz"),
        "annual_holdings_tonnes": (52, "gold_price_usd_oz"),
        "annual_holdings_usd": (61, "gold_price_usd_oz"),
    }
    out: dict[str, list[dict]] = {}
    for name, (start, last_label) in blocks.items():
        data = []
        for r in rows[2:]:
            d = to_date(r[start])
            if d is None:
                continue
            data.append(
                {
                    "date": d.isoformat(),
                    "north_america": num(r[start + 1]),
                    "europe": num(r[start + 2]),
                    "asia": num(r[start + 3]),
                    "other": num(r[start + 4]),
                    last_label: num(r[start + 5]),
                }
            )
        out[name] = data
    return out


def parse_fund_history(meta_list, dates, holdings, demand, flows) -> dict:
    """Per-fund monthly history for drilldown."""
    iso_dates = [d.isoformat() for d in dates]
    funds: dict[str, dict] = {}
    for m in meta_list:
        t = m["ticker"]
        # Trim trailing nones from inactive funds
        h = holdings[t]
        funds[t] = {
            "holdings_tonnes": h,
            "demand_tonnes": demand[t],
            "flows_usd_mn": flows[t],
        }
    return {"dates": iso_dates, "funds": funds}


# ============================================================
# Main
# ============================================================
def main(src: Path | None = None) -> None:
    src = src or latest_xlsx()
    print(f"[parse] Source: {src.name}")
    wb = openpyxl.load_workbook(src, data_only=True, read_only=False)

    # Extract universe once
    meta_list, dates, holdings, demand, flows = extract_fund_universe(wb)
    gp_series = global_series(wb)
    gp_dates = [date.fromisoformat(d) for d in gp_series["dates"]]
    gp_values = gp_series["gold_price_usd_oz"]
    # gold price aligned to same date list
    gp_by_date = dict(zip(gp_dates, gp_values))
    gp_aligned = [gp_by_date.get(d) for d in dates]

    as_of = max(dates)

    funds_out = parse_funds(meta_list, dates, holdings, demand, flows, gp_aligned, as_of)
    regions_out = aggregate(funds_out["funds"], "region")
    countries_out = aggregate(funds_out["funds"], "country")
    top_out = parse_top_movers(funds_out["funds"])
    ts_out = parse_timeseries(wb)
    history_out = parse_fund_history(meta_list, dates, holdings, demand, flows)
    metadata_out = parse_metadata(wb, src, as_of)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUT_DIR / "metadata.json", metadata_out)
    write_json(OUT_DIR / "funds.json", funds_out)
    write_json(OUT_DIR / "regions.json", regions_out)
    write_json(OUT_DIR / "countries.json", countries_out)
    write_json(OUT_DIR / "top_movers.json", top_out)
    write_json(OUT_DIR / "timeseries.json", ts_out)
    write_json(OUT_DIR / "fund_history.json", history_out)
    print("[parse] Done.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", type=Path, help="Path to XLSX (default: latest in data/raw)")
    args = ap.parse_args()
    main(args.src)
